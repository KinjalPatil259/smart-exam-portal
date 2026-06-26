import io
from flask import Blueprint, send_file, session, redirect, url_for, flash, g
from database.connection import db
from models.exam import Exam
from models.attempt import StudentExam, Result
from routes.auth import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from datetime import datetime

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')

@reports_bp.route('/exam/<int:exam_id>/excel')
@login_required
def export_exam_excel(exam_id):
    # Only admins can export full reports
    if session.get('role') != 'admin':
        flash('Unauthorized access.', 'danger')
        return redirect(url_for('student.dashboard'))
        
    exam = db.session.get(Exam, exam_id)
    if not exam:
        flash('Exam session not found.', 'danger')
        return redirect(url_for('admin.reports'))
        
    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Results"
    
    # Title Row
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f"Results Report: {exam.title} ({exam.subject.code})"
    title_cell.font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # Headers
    headers = ["Roll Number", "Student Name", "Department", "Secured Score", "Total Marks", "Percentage", "Grade", "Status", "Submission Time"]
    ws.append([]) # Empty separator row
    ws.append(headers)
    
    # Style Header Row
    header_fill = PatternFill(start_color='1E1B4B', end_color='1E1B4B', fill_type='solid')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='left', vertical='center')
    
    ws.row_dimensions[3].height = 25
    for col_idx in range(1, 10):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        
    # Append Data Rows
    attempts = StudentExam.query.filter_by(exam_id=exam_id, status='submitted').all()
    
    data_font = Font(name='Segoe UI', size=10)
    for index, att in enumerate(attempts):
        res = att.result
        row_data = [
            att.student.roll_number,
            att.student.name,
            att.student.department or 'N/A',
            att.score,
            exam.total_marks,
            f"{res.percentage}%" if res else '--',
            res.grade if res else '--',
            "PASSED" if att.is_passed else "FAILED",
            att.submitted_at.strftime('%Y-%m-%d %H:%M') if att.submitted_at else ''
        ]
        ws.append(row_data)
        
        # Style Data Row
        row_num = index + 4
        ws.row_dimensions[row_num].height = 20
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = data_font
            
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            # Skip merged title row width evaluation
            if cell.row == 1: continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Save to buffer and send file
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"report_exam_{exam_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reports_bp.route('/scorecard/<int:attempt_id>/pdf')
@login_required
def export_scorecard_pdf(attempt_id):
    attempt = db.session.get(StudentExam, attempt_id)
    if not attempt:
        flash('Scorecard session details not found.', 'danger')
        return redirect(url_for('student.dashboard'))
        
    # Security: Only owner or admin can download scorecard
    if session.get('role') != 'admin' and attempt.student_id != g.user.student_profile.id:
        flash('Unauthorized access.', 'danger')
        return redirect(url_for('student.dashboard'))
        
    if attempt.status != 'submitted' or not attempt.result:
        flash('This attempt is not submitted or graded yet.', 'warning')
        return redirect(url_for('student.dashboard'))
        
    exam = attempt.exam
    student = attempt.student
    res = attempt.result
    
    # Generate PDF in Memory Buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#6366f1'),
        alignment=1, # Center
        spaceAfter=5
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor('#475569'),
        alignment=1,
        spaceAfter=25
    )
    
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#1e1b4b'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#334155'),
        leading=14
    )
    
    # Header Branding
    story.append(Paragraph("SMART ONLINE EXAM SYSTEM", title_style))
    story.append(Paragraph("OFFICIAL ASSESSMENT SCORECARD", subtitle_style))
    story.append(Spacer(1, 10))
    
    # Candidate and Exam Details Table
    details_data = [
        [
            Paragraph("<b>Candidate Name:</b>", body_style),
            Paragraph(student.name, body_style),
            Paragraph("<b>Exam Session:</b>", body_style),
            Paragraph(exam.title, body_style)
        ],
        [
            Paragraph("<b>Roll Number:</b>", body_style),
            Paragraph(student.roll_number, body_style),
            Paragraph("<b>Course Subject:</b>", body_style),
            Paragraph(f"{exam.subject.name} ({exam.subject.code})", body_style)
        ],
        [
            Paragraph("<b>Department:</b>", body_style),
            Paragraph(student.department or 'N/A', body_style),
            Paragraph("<b>Date Submitted:</b>", body_style),
            Paragraph(attempt.submitted_at.strftime('%Y-%m-%d %H:%M UTC'), body_style)
        ]
    ]
    
    t_details = Table(details_data, colWidths=[100, 160, 100, 160])
    t_details.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
    ]))
    
    story.append(Paragraph("Candidate & Assessment Details", section_style))
    story.append(t_details)
    story.append(Spacer(1, 20))
    
    # Results Calculations Table
    result_status = "PASSED" if attempt.is_passed else "FAILED"
    status_color = colors.HexColor('#16a34a') if attempt.is_passed else colors.HexColor('#dc2626')
    
    results_data = [
        ["Max Marks", "Marks Secured", "Percentage", "Grade", "Status"],
        [
            str(exam.total_marks),
            str(res.score),
            f"{res.percentage}%",
            res.grade or '--',
            result_status
        ]
    ]
    
    t_results = Table(results_data, colWidths=[100, 100, 100, 100, 120])
    t_results.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e1b4b')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 11),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#f8fafc')),
        ('TEXTCOLOR', (4,1), (4,1), status_color), # Status text color
        ('FONTNAME', (4,1), (4,1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    
    story.append(Paragraph("Performance Metrics Summary", section_style))
    story.append(t_results)
    story.append(Spacer(1, 50))
    
    # Official Signature Footer
    sig_data = [
        [
            Paragraph("System Verification Code: <i>SEC-SYS-" + str(attempt.id) + "</i>", body_style),
            Paragraph("_____________________________<br><b>Authorized Examination Registrar</b>", ParagraphStyle('Sig', parent=body_style, alignment=2))
        ]
    ]
    t_sig = Table(sig_data, colWidths=[260, 260])
    t_sig.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
    ]))
    
    story.append(t_sig)
    
    # Build document
    doc.build(story)
    
    # Return file response
    buffer.seek(0)
    filename = f"scorecard_attempt_{attempt.id}.pdf"
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )
